import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_offer_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Artist Offer"
    
    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    
    # Header styling
    header_font = Font(name='Arial', size=14, bold=True)
    subheader_font = Font(name='Arial', size=12, bold=True)
    normal_font = Font(name='Arial', size=11)
    
    # Border styles
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Title Section
    ws.merge_cells('A1:G1')
    ws['A1'] = "[ARTIST NAME] - Mutually Agreeable"
    ws['A1'].font = header_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:G2')
    ws['A2'] = "Boulder Theatre Benefit Concert ( Sweet Jam Benefit for Asheville )"
    ws['A2'].font = subheader_font
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:G3')
    ws['A3'] = "Headliner Offer"
    ws['A3'].font = subheader_font
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Guarantee
    ws.merge_cells('A4:B4')
    ws['A4'] = "Guarantee:"
    ws['A4'].font = Font(bold=True)
    ws['C4'] = "$[AMOUNT]"
    
    # Event Details Section
    ws.merge_cells('A6:G6')
    ws['A6'] = "Event Details"
    ws['A6'].font = subheader_font
    ws['A6'].fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    
    # Event Description
    ws.merge_cells('A7:G9')
    ws['A7'] = "This invitation is to participate in the Sweet Jam Benefit for Asheville, a charitable concert organized to support Asheville, North Carolina in the aftermath of Hurricane Helena. The event will feature renowned musicians coming together for an evening of music and community support at the historic Boulder Theatre."
    ws['A7'].alignment = Alignment(wrap_text=True)
    
    # Ticket Scaling Section
    ws.merge_cells('A11:G11')
    ws['A11'] = "Ticket Scaling"
    ws['A11'].font = subheader_font
    ws['A11'].fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    
    # Ticket Table Headers
    headers = ['Type', 'Allotment', 'Comps', 'Sellable', 'Ticket Price', 'Per Ticket Fees', 'Net Price', 'Adjusted Gross Potential']
    for col, header in enumerate(headers, 1):
        ws.cell(row=12, column=col).value = header
        ws.cell(row=12, column=col).font = Font(bold=True)
        ws.cell(row=12, column=col).border = thin_border
    
    # Contact Information
    ws.merge_cells('A20:G20')
    ws['A20'] = "Contacts"
    ws['A20'].font = subheader_font
    ws['A20'].fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    
    contacts = [
        ('Email:', 'JamcamAnnemarie@gmail.com'),
        ('Phone:', '303.330.8814'),
        ('Website:', 'sweetjambenefit.org')
    ]
    
    current_row = 21
    for label, value in contacts:
        ws.cell(row=current_row, column=1).value = label
        ws.cell(row=current_row, column=2).value = value
        current_row += 1
    
    # Deal Terms
    ws.merge_cells(f'A{current_row+1}:G{current_row+1}')
    ws[f'A{current_row+1}'] = "Deal Terms"
    ws[f'A{current_row+1}'].font = subheader_font
    ws[f'A{current_row+1}'].fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    
    # Add deal terms with proper formatting
    deal_terms = [
        ('Ages:', '16+'),
        ('COMPS:', 'TBA'),
        ('ALL INCLUSIVE:', 'Offer includes backline, VIP ground transportation, and lodging. Rider requirements provided per budget and advance.'),
        ('SEND CONTRACTS TO:', 'JamcamAnnemarie@gmail.com'),
        ('SEND TICKET COUNTS TO:', 'JamcamAnnemarie@gmail.com'),
        ('EXCLUSIVITY:', 'Radius clause does not apply for this benefit event.'),
        ('DEPOSITS:', 'Deposits of [X]% will be sent no sooner than [X] days out from the show. Balance will be settled night of show via check or the next day by ACH.'),
        ('MARKETING:', 'Artist agrees to support event marketing through their established in-house marketing divisions, including co-branding opportunities and promotional support across their platforms.'),
        ('MERCHANDISE:', 'Artist merchandise percentage currently being negotiated // Artist sells.'),
        ('BACKLINE:', 'Backline will be provided.'),
        ('EXPIRATION:', 'Offer expires 14 days after received.')
    ]
    
    current_row += 2
    for label, value in deal_terms:
        ws.cell(row=current_row, column=1).value = label
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.merge_cells(f'B{current_row}:G{current_row}')
        ws.cell(row=current_row, column=2).value = value
        current_row += 1
    
    return wb

# Create the template
template = create_offer_template()

# Save the template
template.save('sweet_jam_benefit_offer_template.xlsx')

# Debug: Print some information about the workbook
print(f"Workbook created with {len(template.worksheets)} worksheet(s)")
print(f"Active worksheet title: {template.active.title}")
print(f"Number of rows in the active worksheet: {template.active.max_row}")
print(f"Number of columns in the active worksheet: {template.active.max_column}")

# Debug: Print the content of some cells
print("\nCell contents:")
for row in range(1, 5):
    for col in range(1, 4):
        cell = template.active.cell(row=row, column=col)
        print(f"Cell {cell.coordinate}: {cell.value}")

# Debug: Check if merged cells are correct
print("\nMerged cell ranges:")
for merged_range in template.active.merged_cells.ranges:
    print(merged_range)

# Debug: Check column widths
print("\nColumn widths:")
for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
    print(f"Column {col}: {template.active.column_dimensions[col].width}")
